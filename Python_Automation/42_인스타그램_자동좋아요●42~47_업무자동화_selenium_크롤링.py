# 한글의 자음 모음 분리 현상 해결 : unicodedata.normalize("NFC", "한글 문자")

from selenium import webdriver
import time
import random
import os
import openpyxl
import unicodedata

#인스타그램 엑셀파일 만들기 & 열기
if not os.path.exists("./인스타그램.xlsx"):
    book = openpyxl.Workbook()
    book.save("./인스타그램.xlsx")
book = openpyxl.load_workbook("./인스타그램.xlsx")
sheet = book.active #sheet = book["Sheet1"]

#검색할 hash_tag 입력
hash_tag = input("해시태그 입력 >>")

#브라우저 열기
browser = webdriver.Chrome("./chromedriver.exe") #chromedriver 파일을 불러옴
browser.get("https://www.instagram.com/accounts/login/?source=auth_switcher")
time.sleep(3)

#로그인
id = browser.find_element_by_name("username") #css선택자 없이도 name 속성값만 전달하면 요소를 가져옴
id.send_keys("asdf7898897@gmail.com")
pw = browser.find_element_by_name("password")
pw.send_keys("qwer1232231!")
browser.find_element_by_css_selector("div.Igw0E.IwRSH.eGOV_._4EzTm.bkEs3.CovQj.jKUp7.DhRcB").click()
time.sleep(4)

#hash_tag 검색
url = "https://www.instagram.com/explore/tags/{}/".format(hash_tag)
browser.get(url)
time.sleep(7)

#첫번째 사진 클릭
browser.find_element_by_css_selector("div._9AhH0").click()
time.sleep(3)

row_num = 1

#자동 좋아요 실행 & 닉네임과 내용 엑셀파일에 저장
while True:
    # 만약에 좋아요가 안 눌려져 있다면?(aria-label의 속성값이 "좋아요"라면?)
    #     좋아요 누르고
    #     다음 사진 넘어가고
    # 만약 좋아요가 눌려져 있다면?(aria-label의 속성값이 "좋아요 취소"라면?)
    #     다음 사진 넘어가고

    #요소 가져오기
    like = browser.find_element_by_css_selector("span.fr66n > button.wpO6b > div.QBdPU svg._8-yf5")
    likevalue = like.get_attribute("aria-label")
    next = browser.find_element_by_css_selector("a._65Bje.coreSpriteRightPaginationArrow")
    nick_name = browser.find_element_by_css_selector("a.sqdOP.yWX7d._8A5w5.ZIAjV ") #a.sqdOP.yWX7d._8A5w5.ZIAjV에 해당하는 요소가 많지만, select를 사용하면 제일 위에 것만 불러오므로 상관없음
    content = browser.find_element_by_css_selector("div.C4VMK > span") #div.C4VMK > span에 해당하는 요소가 많지만, select를 사용하면 제일 위에 것만 불러오므로 상관없음
    content_normalize = unicodedata.normalize("NFC", content.text) # 한글의 자음 모음 분리 현상 해결

    # 닉네임과 내용 엑셀파일에 저장
    sheet.cell(row=row_num, column=1).value = nick_name.text
    sheet.cell(row=row_num, column=2).value = content_normalize
    row_num += 1
    book.save("./인스타그램.xlsx")

    #자동 좋아요 실행
    if likevalue == "좋아요":
        like.click() #likevalue.click() : 속성값을 클릭하는 것이 되므로 안됨 / like.click() : css 선택자 전체를 누르는 것이므로 좋아요 버튼을 클릭하는 것이 됨
        time.sleep(random.randint(2,5) + random.random())
        next.click()
        time.sleep(random.randint(2,5) + random.random())
    elif likevalue == "좋아요 취소":
        next.click()
        time.sleep(random.randint(2,5) + random.random())