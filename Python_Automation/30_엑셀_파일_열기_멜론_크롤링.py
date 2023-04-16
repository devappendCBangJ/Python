#View > Tool Windows > Terminal에서 Terminal창 >> pip install openpyxl 입력
#<엑셀 모듈>
# 0. 모듈 불러오기 : import openpyxl
# 0. 엑셀에 넣을 수 있는 이미지로 변환 시켜주는 모듈 : from openpyxl.drawing.image import Image
# 1. 엑셀 파일 생성 : openpyxl.Workbook()
# 2. 엑셀 파일 저장 : openpyxl.Workbook.save("파일 경로/파일명.xlsx")
# 3. 엑셀파일 불러오기 : openpyxl.load_workbook("파일 경로/파일명.xlsx")
# 4. 시트 사용 : openpyxl.load_workbook["Sheet1"] #엑셀 파일에서 Sheet1을 사용하겠다
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active #엑셀 파일 열었을 때 자동으로 열리는 그 시트를 사용하겠다
# 5. 열 너비 지정 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.coulumn_dimensions["A..."].width = 열의 너비
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["A"].width = 15 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["B"].width = 50.38 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["C"].width = 79.13 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["D"].width = 56.50 #해당 열의 너비 지정
# 6. 행 높이 지정 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.row_dimensions[row_num].height = 95  # 해당 행의 높이 지정. 높이는 for문에서 돌려서 하나하나 바뀌도록 설정
# 7. 셀에 이미지 넣기 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.add_image(Image("파일 경로/파일명.png"), "A{}".format(row_num) 형식의 셀 이름)
# 8. 셀에 텍스트 넣기 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.cell(row=row_num, column=column_num).value = 텍스트
# 9. 엑셀 파일 저장 : openpyxl.load_workbook("파일 경로/파일명.xlsx").save("파일 경로/파일명.xlsx")

#ctrl+d : 같은 내용을 다음 행에 복사 붙여 넣기
# 요소 안에 또다른 요소가 들어있을 경우에는 none으로 출력되므로 .string이 아닌 .text사용

import urllib.request as req
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl.drawing.image import Image

#엑셀 파일 생성
if not os.path.exists("./멜론음원차트.xlsx"):
    book = openpyxl.Workbook()
    book.save("./멜론음원차트.xlsx")

#크롤링
headers = req.Request("https://www.melon.com/chart/", headers = {"User-Agent":"Mozilla/5.0"})
code = req.urlopen(headers)
soup = BeautifulSoup(code, "html.parser")
title = soup.select("div.ellipsis.rank01 a")
singer = soup.select("div.ellipsis.rank02 > span.checkEllipsis")
album = soup.select("div.ellipsis.rank03 > a")
image = soup.select("a.image_typeAll > img")

# soup.select("div.ellipsis.rank02 > a") #웹 검사창에서 ctrl+f로 검색하면 100개 이상이 떠버리는데,
# - 이는 top100에서 노래별로 가수를 100개 보내주는게 아니라, 1개 노래에 여러 명의 가수가 있으면, 따로따로 보내줘서 그런 것임
# - 이를 해결하려면 하나의 노래에서 여러 명의 가수를 1개로 취급하는 속성명과 속성값을 찾아서 이용

#
book = openpyxl.load_workbook("./멜론음원차트.xlsx")
sheet = book.active
sheet.column_dimensions["A"].width = 15 #해당 열의 너비 지정
sheet.column_dimensions["B"].width = 50.38 #해당 열의 너비 지정
sheet.column_dimensions["C"].width = 79.13 #해당 열의 너비 지정
sheet.column_dimensions["D"].width = 56.50 #해당 열의 너비 지정
row_num = 1

#이미지 저장할 폴더 만들기
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")

for i in range(len(title)):
    req.urlretrieve(image[i].attrs["src"], "./멜론이미지/{}.png".format(row_num))
    print(title[i].string, singer[i].text, album[i].string, image[i].attrs["src"])
    #.string했는데 none이 출력되면 .text사용
    img_for_excel = Image("./멜론이미지/{}.png".format(row_num))
    sheet.add_image(img_for_excel, "A{}".format(row_num))
    sheet.row_dimensions[row_num].height = 95
    sheet.cell(row=row_num, column=2).value = title[i].string
    sheet.cell(row=row_num, column=3).value = singer[i].text
    sheet.cell(row=row_num, column=4).value = album[i].string
    book.save("./멜론음원차트.xlsx")
    row_num += 1
    #.value 전까지는 셀을 선택만 한 상태
    #.value까지 하면 셀에 무언가를 쓸 수 있는 상태