from bs4 import BeautifulSoup
import urllib.request as req
import openpyxl # 터미널 창에 pip install openpyxl
import os
from openpyxl.drawing.image import Image

# 엑셀 파일 존재하는 확인
if not os.path.exists("./멜론음원차트.xlsx"):
    book = openpyxl.Workbook()
    book.save("./멜론음원차트.xlsx")


headers = req.Request("https://www.melon.com/chart/", headers={"User-Agent":"Mozilla/5.0"})
code = req.urlopen(headers)
soup = BeautifulSoup(code, "html.parser")
title = soup.select("div.ellipsis.rank01 a")
name = soup.select("div.ellipsis.rank02 > span.checkEllipsis")
album = soup.select("div.ellipsis.rank03 > a")
image = soup.select("a.image_typeAll > img")

book = openpyxl.load_workbook("./멜론음원차트.xlsx")
# sheet = book["Sheet1"]
sheet = book.active # 자동으로 열리는 그 시트를 사용하겠다.
sheet.column_dimensions["A"].width = "15"
sheet.column_dimensions["B"].width = "50"
sheet.column_dimensions["C"].width = "30"
sheet.column_dimensions["D"].width = "35"

row_num = 1
# 이미지 저장할 폴더 만들기
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")
for i in range(len(title)):
    req.urlretrieve(image[i].attrs["src"], "./멜론이미지/{}.png".format(row_num))
    print(title[i].string, name[i].text, album[i].string, image[i].attrs["src"])
    img_for_excel = Image("./멜론이미지/{}.png".format(row_num))
    sheet.row_dimensions[row_num].height = 95
    sheet.add_image(img_for_excel, "A{}".format(row_num))
    sheet.cell(row=row_num, column=2).value = title[i].string
    sheet.cell(row=row_num, column=3).value = name[i].text
    sheet.cell(row=row_num, column=4).value = album[i].string
    book.save("./멜론음원차트.xlsx")
    row_num += 1







