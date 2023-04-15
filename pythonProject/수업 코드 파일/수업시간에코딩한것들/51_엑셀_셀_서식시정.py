import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side

book = openpyxl.load_workbook("./엑셀데이터/2017년_광고비_total.xlsx")
sheet = book.active

# 서식 만들기
user_font = Font(name="맑은 고딕", size=12, bold=True)
user_alignment = Alignment(horizontal="center")
orange_color = PatternFill(patternType="solid", fgColor=Color("FE9A2E"))
gray_color = PatternFill(patternType="solid", fgColor=Color("BDBDBD"))
user_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"), bottom=Side(style="thin"))

sheet["A14"].value = "합계"
sheet["A14"].font = user_font
sheet["A14"].alignment = user_alignment
sheet["A14"].fill = orange_color
sheet["A14"].border = user_border

sheet["B14"].value = "=SUM(B2:B13)"
sheet["C14"].value = "=SUM(C2:C13)"
sheet["D14"].value = "=SUM(D2:D13)"

# 여러 셀 서식 지정 한번에
for row in sheet["B14:D14"]:
    for cell in row:
        cell.font = user_font
        cell.alignment = user_alignment
        cell.fill = orange_color
        cell.border = user_border

for row in sheet["B2:D13"]:
    for cell in row:
        cell.font = user_font
        cell.alignment = user_alignment
        cell.fill = gray_color
        cell.border = user_border

book.save("./엑셀데이터/2017년_광고비_total_서식시정.xlsx")





