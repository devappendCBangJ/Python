from selenium import webdriver
import time
import random
import openpyxl
import os
import unicodedata

if not os.path.exists("./인스타그램.xlsx"):
    book = openpyxl.Workbook()
    book.save("./인스타그램.xlsx")

book = openpyxl.load_workbook("./인스타그램.xlsx")
# sheet = book["Sheet1"]
sheet = book.active

hash_tag = input("해시태그 입력 >> ")
browser = webdriver.Chrome("./chromedriver")
browser.get("https://www.instagram.com/accounts/login/?source=auth_switcher")
time.sleep(3)

id = browser.find_element_by_name("username")
id.send_keys("아이디")
pw = browser.find_element_by_name("password")
pw.send_keys("비밀번호")
browser.find_element_by_css_selector("div.Igw0E.IwRSH.eGOV_._4EzTm.bkEs3.CovQj.jKUp7.DhRcB").click()
time.sleep(4)
# 해시태그 검색하기
url = "https://www.instagram.com/explore/tags/{}/".format(hash_tag)
browser.get(url)
time.sleep(7)
# 첫번째 사진 클릭하기
browser.find_element_by_css_selector("div._9AhH0").click()
time.sleep(3)
# 자동 좋아요 동작시키기
row_num = 1
while True:
    like = browser.find_element_by_css_selector("section.ltpMr.Slqrh button.wpO6b svg._8-yf5")
    value = like.get_attribute("aria-label")
    next = browser.find_element_by_css_selector("a._65Bje.coreSpriteRightPaginationArrow")
    ### 크롤링 ###
    nick_name = browser.find_element_by_css_selector("a.sqdOP.yWX7d._8A5w5.ZIAjV")
    content = browser.find_element_by_css_selector("div.C4VMK > span")
    content_normalize = unicodedata.normalize("NFC", content.text)  # 한글의 자음 모음 분리 현상을 해결.
    sheet.cell(row=row_num, column=1).value = nick_name.text
    sheet.cell(row=row_num, column=2).value = content_normalize
    row_num += 1
    book.save("./인스타그램.xlsx")
    if value == "좋아요":
        like.click()
        time.sleep(random.randint(2,5) + random.random())
        next.click()
        time.sleep(random.randint(2,5) + random.random())
    elif value == "좋아요 취소":
        next.click()
        time.sleep(random.randint(2,5) + random.random())






