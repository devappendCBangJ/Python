import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
from openpyxl.chart import Reference, Series, LineChart, BarChart, PieChart
import pandas as pd

df = pd.read_excel("./엑셀데이터/아파트분양가격.xlsx")
locations = df["지역명"].unique()
for loc in locations:
    writer = pd.ExcelWriter("./엑셀데이터/아파트분양가격_{}.xlsx".format(loc))
    df_location = df[df["지역명"] == loc]
    years = df_location["연도"].unique()
    for y in years:
        df_year = df_location[df_location["연도"] == y]
        df_result = df_year.sort_values(by="월", ascending=True)
        df_result.to_excel(writer, sheet_name="{}년".format(y), index=None)
    writer.save()
    print("{} 지역 엑셀파일 분리 완료".format(loc))

###### 엑셀 셀 서식 만들기
user_font = Font(name="맑은 고딕", size=12, bold=True)
user_alignment = Alignment(horizontal="center")
orange_color = PatternFill(patternType="solid", fgColor=Color("FE9A2E"))
gray_color = PatternFill(patternType="solid", fgColor=Color("BDBDBD"))
user_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"), bottom=Side(style="thin"))
#######

for loc in locations:
    book = openpyxl.load_workbook("./엑셀데이터/아파트분양가격_{}.xlsx".format(loc))
    for sheetname in book.sheetnames:
        sheet = book[sheetname]
        # 데이터 헤더에 셀 서식 지정
        for row in sheet["A1:E1"]:
            for cell in row:
                cell.font = user_font
                cell.alignment = user_alignment
                cell.fill = orange_color
                cell.border = user_border
        # 나머지 셀 부분에 서식 시정
        for row in sheet["A2:E{}".format(sheet.max_row)]:
            for cell in row:
                cell.alignment = user_alignment
                cell.fill = gray_color
                cell.border = user_border

        # 차트 그리기
        chart = BarChart()
        chart.title = "{} 지역 {} 년도 아파트 분양가격".format(loc, sheetname)
        value = Reference(sheet, range_string="{}!E2:E{}".format(sheetname, sheet.max_row))
        value_series = Series(value, title="분양가격")
        chart.append(value_series)
        sheet.add_chart(chart, "G1")
    book.save("./엑셀데이터/아파트분양가격_{}.xlsx".format(loc))
    print("{} 지역 엑셀 서식 지정 완료".format(loc))

