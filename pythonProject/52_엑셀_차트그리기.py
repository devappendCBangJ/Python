#View > Tool Windows > Terminal에서 Terminal창 >> pip install openpyxl 입력

#<엑셀 모듈>
# 0. 모듈 불러오기 : import openpyxl
# 0. 엑셀 서식 변경 모듈 : from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
# 0. 엑셀 차트 그리기 모듈 : from openpyxl.chart import Reference, Series, LineChart, BarChart, PieChart
# 0. 엑셀에 넣을 수 있는 이미지로 변환 시켜주는 모듈 : from openpyxl.drawing.image import Image
# 1. 엑셀 파일 생성 : openpyxl.Workbook()
# 2. 엑셀 파일 저장 : openpyxl.Workbook.save("파일 경로/파일명.xlsx")
# 3. 엑셀파일 불러오기 : openpyxl.load_workbook("파일 경로/파일명.xlsx")
# 4. 시트 사용 : openpyxl.load_workbook["Sheet1"] #엑셀 파일에서 Sheet1을 사용하겠다
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active #엑셀 파일 열었을 때 자동으로 열리는 그 시트를 사용하겠다
# 5. 열 너비 지정 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.coulumn_dimensions["A..."].width = 열의 너비
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["A"].width = 15 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["B"].width = 50.38 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["C"].width = 79.13 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["D"].width = 56.50 #해당 열의 너비 지정
# 6. 행 높이 지정 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.row_dimensions[row_num].height = 95  # 해당 행의 높이 지정. 높이는 for문에서 돌려서 하나하나 바뀌도록 설정
# 7. 셀에 이미지 넣기 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.add_image(Image("파일 경로/파일명.png"), "A{}".format(row_num) 형식의 셀 이름)
# 8. 셀에 텍스트 넣기1 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.cell(row=row_num, column=column_num).value = 텍스트
# 8. 셀에 텍스트 넣기2 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].value = 텍스트
# 9. 셀 서식 - 폰트 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].font = Font(name="맑은 고딕", size=12, bold=True)
# 9. 셀 서식 - 정렬 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].alignment = Alignment(horizontal="center")
# 9. 셀 서식 - 글자색 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].fill = PatternFill(patternType="solid", fgColor=Color("AACBE8")) # 샵은 제외하고 컬러코드 넣기
# 9. 셀 서식 - 셀 테두리 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
# 10. 셀에 차트 넣기 :
    # 1) 차트 도화지 생성 : chart = LineChart() or BarChart() or PieChart()...
    # 2) 차트 제목 설정 : chart.title = "차트 제목"
    # 3) 차트로 사용될 값 지정 :
    # for j in ["C", "D", "E", "F", "G"]:
        # reference = Reference(openpyxl.load_workbook("파일 경로/파일명.xlsx").active, range_string="Sheet1!{}2:{}13".format(j, j))
        # series = Series(reference, title=openpyxl.load_workbook("파일 경로/파일명.xlsx").active["{}1".format(j)].value)
        # chart.append(series)
    # 4) 셀에 차트 넣기 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.add_chart(chart, "J1")
# 11. 엑셀 파일 저장 : openpyxl.load_workbook("파일 경로/파일명.xlsx").save("파일 경로/파일명.xlsx") # 확장자 필수

# 여러 셀 서식 지정
# for row in sheet["B14:D14"]: # row엔 [B14:D14]이 들어있음
#     for cell in row: # cell엔 B14, C14, D14가 들어있음
#         cell.font = user_font
#         cell.alignment = user_alignment
#         cell.fill = blue_color
#         cell.border = user_border

# 차트에 값 대입
# for j in ["C", "D", "E", "F", "G"]:
#     value = Reference(sheet, range_string="Sheet1!{}2:{}13".format(j, j))
#     value_series = Series(value, title=sheet["{}1".format(j)].value)
#     chart.append(value_series)

import openpyxl
from openpyxl.chart import Reference, Series, LineChart, BarChart, PieChart

company = ["LG전자", "삼성전자", "현대자동차"]

for i in company:
    # 엑셀 불러오기
    book = openpyxl.load_workbook("./엑셀데이터/2017년_광고비_{}.xlsx".format(i))
    sheet = book.active

    # 차트 도화지 생성
    chart = LineChart() # 라인 차트를 쓰고 싶다면 LineChart, 막대 차트를 쓰고 싶다면 BarChart, 파이 차트를 쓰고 싶다면 PieChart
    # 차트 제목
    chart.title = "{} 월별 광고".format(i)
    # 차트에 값 대입
    for j in ["C", "D", "E", "F", "G"]:
        value = Reference(sheet, range_string="Sheet1!{}2:{}13".format(j, j))
        value_series = Series(value, title=sheet["{}1".format(j)].value)
        chart.append(value_series)
    # 차트 시각화
    sheet.add_chart(chart, "J1")
    book.save("./엑셀데이터/2017년_광고비_{}_차트.xlsx".format(i)) #파일명 쓸 때 확장자 빼먹으면 컴퓨터가 인식을 못함