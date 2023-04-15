#View > Tool Windows > Terminal에서 Terminal창 >> pip install openpyxl 입력

#<엑셀 모듈>
# 0. 모듈 불러오기 : import openpyxl
# 0. 엑셀 서식 변경 모듈 : from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
# 0. 엑셀에 넣을 수 있는 이미지로 변환 시켜주는 모듈 : from openpyxl.drawing.image import Image
# 1. 엑셀 파일 생성 : openpyxl.Workbook()
# 2. 엑셀 파일 저장 : openpyxl.Workbook.save("파일 경로/파일명.xlsx")
# 3. 엑셀파일 불러오기 : openpyxl.load_workbook("파일 경로/파일명.xlsx")
# 4. 시트 사용 : openpyxl.load_workbook["Sheet1"] #엑셀 파일에서 Sheet1을 사용하겠다
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active #엑셀 파일 열었을 때 자동으로 열리는 그 시트를 사용하겠다
# 5. 열 너비 지정 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.coulumn_dimensions["A..."].width = 열의 너비
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["A"].width = 15 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["B"].width = 50.38 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["C"].width = 79.13 #해당 열의 너비 지정
#               openpyxl.load_workbook("파일 경로/파일명.xlsx").active.column_dimensions["D"].width = 56.50 #해당 열의 너비 지정
# 6. 행 높이 지정 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.row_dimensions[row_num].height = 95  # 해당 행의 높이 지정. 높이는 for문에서 돌려서 하나하나 바뀌도록 설정
# 7. 셀에 이미지 넣기 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.add_image(Image("파일 경로/파일명.png"), "A{}".format(row_num) 형식의 셀 이름)
# 8. 셀에 텍스트 넣기1 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active.cell(row=row_num, column=column_num).value = 텍스트
# 8. 셀에 텍스트 넣기2 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].value = 텍스트
# 9. 셀 서식 - 폰트 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].font = Font(name="맑은 고딕", size=12, bold=True)
# 9. 셀 서식 - 정렬 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].alignment = Alignment(horizontal="center")
# 9. 셀 서식 - 글자색 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].fill = PatternFill(patternType="solid", fgColor=Color("AACBE8")) # 샵은 제외하고 컬러코드 넣기
# 9. 셀 서식 - 셀 테두리 : openpyxl.load_workbook("파일 경로/파일명.xlsx").active["A14"].border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
# 10. 엑셀 파일 저장 : openpyxl.load_workbook("파일 경로/파일명.xlsx").save("파일 경로/파일명.xlsx")

# 엑셀파일에서 시트 이름들을 불러옴 : openpyxl.load_workbook("파일 위치/파일명").sheetnames
# 최대 행까지 범위 지정 : for row in openpyxl.load_workbook("파일 위치/파일명")["시트 이름"]["A1:E{}".format(openpyxl.load_workbook("파일 위치/파일명")["시트 이름"].max_row)]:

# 여러 셀 서식 지정
# for row in sheet["B14:D14"]: # sheet["B14:D14"]에 [[B14, C14, D14]]가 있으므로 row엔 [B14, C14, D14]가 들어있음
#     for cell in row: # cell는 B14, C14, D14임
#         cell.font = user_font
#         cell.alignment = user_alignment
#         cell.fill = blue_color
#         cell.border = user_border

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side

# 엑셀 불러오기
book = openpyxl.load_workbook("./엑셀데이터/2017년_광고비_total.xlsx")
sheet = book.active

# 서식 만들기
user_font = Font(name="맑은 고딕", size=12, bold=True)
user_alignment = Alignment(horizontal="center")
blue_color = PatternFill(patternType="solid", fgColor=Color("AACBE8"))
yellow_color = PatternFill(patternType="solid", fgColor=Color("FAF193"))
user_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

# 셀 값 & 서식 지정
sheet["A14"].value = "합계" # sheet.cell(row = 14, column = 1)와 같음
sheet["A14"].font = user_font
sheet["A14"].alignment = user_alignment
sheet["A14"].fill = blue_color
sheet["A14"].border = user_border

sheet["B14"].value = "=SUM(B2:B13)"
sheet["C14"].value = "=SUM(C2:C13)"
sheet["D14"].value = "=SUM(D2:D13)"

# 여러 셀 서식 지정
for row in sheet["B14:D14"]:
    for cell in row:
        cell.font = user_font
        cell.alignment = user_alignment
        cell.fill = blue_color
        cell.border = user_border

for row in sheet["B2:D13"]:
    for cell in row:
        cell.font = user_font
        cell.alignment = user_alignment
        cell.fill = yellow_color
        cell.border = user_border

book.save("./엑셀데이터/2017년_광고비_total_서식지정.xlsx")