# 열에서 중복값 제외 고유한 값 리스트화 : pd.read_excel("파일 위치/파일명")["열 제목"].unique()
# 엑셀파일에서 시트 이름들을 불러옴 : openpyxl.load_workbook("파일 위치/파일명").sheetnames
# 최대 행까지 범위 지정 : for row in openpyxl.load_workbook("파일 위치/파일명")["시트 이름"]["A1:E{}".format(openpyxl.load_workbook("파일 위치/파일명")["시트 이름"].max_row)]:
# Ctrl + 원하는 함수 or 변수 클릭 >> 함수가 정의된 위치 or 변수가 정의된 위치로 이동

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference, Series

# 엑셀 불러오기
df = pd.read_excel("./엑셀데이터/아파트분양가격.xlsx")
print(df["지역명"].unique()) # 확인용 코드

# 조건에 맞는 값만 뽑은 후 저장
locations = df["지역명"].unique()
for loc in locations:
    writer = pd.ExcelWriter("./엑셀데이터/아파트분양가격_{}.xlsx".format(loc))
    df_location = df[df["지역명"] == loc]
    years = df_location["연도"].unique()
    for y in years:
        df_year = df_location[df_location["연도"] == y]
        df_result = df_year.sort_values(by="월", ascending=True)
        df_result.to_excel(writer, sheet_name="{}년".format(y), index=None)
    writer.save()
    print("{} 지역 엑셀파일 분리 완료".format(loc))

# 서식 만들기
user_font = Font(name="맑은 고딕", size=12, bold=True)
user_alignment = Alignment(horizontal="center")
blue_color = PatternFill(patternType="solid", fgColor=Color("AACBE8"))
yellow_color = PatternFill(patternType="solid", fgColor=Color("FAF193"))
user_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

# 엑셀 불러오기
for loc in locations:
    book = openpyxl.load_workbook("./엑셀데이터/아파트분양가격_{}.xlsx".format(loc))
    for sheetname in book.sheetnames:
        sheet = book[sheetname]
        
        # 데이터 헤더에 셀 서식 지정
        for row in sheet["A1:E1"]:
            for cell in row:
                cell.font = user_font
                cell.alignment = user_alignment
                cell.fill = yellow_color
                cell.border = user_border
        
        # 나머지 셀 부분에 서식 지정
        for row in sheet["A2:E{}".format(sheet.max_row)]:
            for cell in row:
                cell.alignment = user_alignment
                cell.fill = blue_color
                cell.border = user_border

        # 차트 그리기
        chart = BarChart()
        chart.title = "{}지역 {}년도 아파트 분양가격".format(loc, sheetname)
        value = Reference(sheet, range_string="{}!E2:E{}".format(sheetname, sheet.max_row))
        value_series = Series(value, title="분양가격")
        chart.append(value_series)
        sheet.add_chart(chart, "G1")

    # 엑셀 저장
    book.save("./엑셀데이터/아파트분양가격_{}.xlsx".format(loc))
    print("{} 지역 엑셀 서식 지정 완료".format(loc))